# -*- coding: utf-8 -*-
"""
Created on Mon Jan 14 00:42:11 2019

@author: XIAN
"""

#molden

import re
from openpyxl import load_workbook
import os


path = '/Users/XIAN/Documents/xian fyp/new molden/temp'
lw = load_workbook('/Users/XIAN/Documents/xian fyp/temp.xlsx')
lww = load_workbook('/Users/XIAN/Documents/xian fyp/1st batch.xlsx')

sheet = lw.get_active_sheet()
sheet2 = lww.get_active_sheet()
a = 1

for filename in os.listdir(path):
    if filename.endswith(".molden"):
        print(filename) 
        name = str(re.search('\d+', filename).group()) 
        print(name)                      
            
        i=0
        ii = 0
        ii2 = 0
           
        with open(os.path.join(path, filename), 'r+') as f:
            for line in f.readlines():
                i=i+1 #enumerate number of line
                    
                mystring = str(line)
                sheet.cell(row=i, column=1).value=mystring #insert into temp.xls file first line
                if re.search('Occup= 2.0', line):   
                    ii=i
                        
                if re.search(' Occup= 0.0', line):
                    ii2=i
                    break                        
                        
        lw.save('/Users/XIAN/Documents/xian fyp/temp.xlsx')    
        mystr = sheet.cell(row=ii-2, column=1).value #locate value needed
        ans = re.findall("-\d+\.\d+",mystr) # extract the value        
        for i in ans:
            myfloat = float(i)
        print (ans)   #print the value list
            
        mystr = sheet.cell(row=ii2-2, column=1).value #locate value needed
        ans2 = re.findall("-\d+\.\d+",mystr) # extract the value        
        for i in ans2:
            myfloat2 = float(i) 
        print (ans2)   #print the value list
            
        a = a+1
        sheet2.cell(row=a, column =1).value= name
            
        sheet2.cell(row=a, column =2).value= myfloat
        sheet2.cell(row=a, column =3).value= myfloat2
        lww.save('/Users/XIAN/Documents/xian fyp/1st batch.xlsx')        
         
            
        old1 = os.path.join(path, filename)
        new1 = '/Users/XIAN/Documents/xian fyp/new molden/complete/'+ str(filename)
        os.rename(old1,new1)
        
        print('done')
    
    